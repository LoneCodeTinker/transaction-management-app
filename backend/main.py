from fastapi import FastAPI, HTTPException, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from typing import List, Optional
import openpyxl
import os
from datetime import date, datetime
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile
import shutil

app = FastAPI()

# Allow frontend (localhost:5173) to access backend
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "*"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"]
)

EXCEL_FILE = "transactions.xlsx"
SHEET_NAMES = {"sales": "Sales", "received": "Received", "purchases": "Purchases", "expenses": "Expenses"}

class Transaction(BaseModel):
    type: str  # sales, purchases, expenses, received
    name: str  # vendor or customer
    date: date
    description: Optional[str] = None
    reference: Optional[str] = None
    amount: float
    vat: float = 0
    total: float
    method: Optional[str] = None  # Only for received
    actions: Optional[List[str]] = None
    done: bool = False


def get_or_create_workbook():
    # Try to load workbook, handle corruption
    try:
        if not os.path.exists(EXCEL_FILE):
            wb = openpyxl.Workbook()
            for sheet in SHEET_NAMES.values():
                wb.create_sheet(sheet)
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            wb.save(EXCEL_FILE)
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except (BadZipFile, InvalidFileException):
        # Corrupted file: backup and recreate
        backup_name = f"transactions_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        shutil.move(EXCEL_FILE, backup_name)
        wb = openpyxl.Workbook()
        for sheet in SHEET_NAMES.values():
            wb.create_sheet(sheet)
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        wb.save(EXCEL_FILE)
        wb = openpyxl.load_workbook(EXCEL_FILE)
    # Ensure all required sheets exist and have correct headers
    changed = False
    for key, sheet_name in SHEET_NAMES.items():
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
            changed = True
        sheet = wb[sheet_name]
        # Fix: Always ensure headers are in the first row, and remove any extra header rows
        if sheet.max_row == 0 or all(cell.value is None for cell in sheet[1]):
            if key == 'received':
                sheet.append(["Name", "Date", "Amount", "Notes", "Method", "Actions", "Done"])
            else:
                sheet.append(["Name", "Date", "Description", "Reference", "Amount", "VAT", "Total", "Actions", "Done"])
            changed = True
        # Remove duplicate header rows (keep only the first row as header)
        header_values = [cell.value for cell in sheet[1]]
        rows_to_delete = []
        for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
            if [cell.value for cell in row] == header_values:
                rows_to_delete.append(i)
        for i in reversed(rows_to_delete):
            sheet.delete_rows(i)
        # Also, if the first row is empty, remove it
        if all(cell.value is None for cell in sheet[1]):
            sheet.delete_rows(1)
    if changed:
        wb.save(EXCEL_FILE)
    return wb


def save_transaction_to_excel(tx: Transaction):
    wb = get_or_create_workbook()
    sheet = wb[SHEET_NAMES[tx.type]]
    if sheet.max_row == 1:
        if tx.type == 'received':
            sheet.append([
                "Name", "Date", "Amount", "Notes", "Method", "Actions", "Done"
            ])
        else:
            sheet.append([
                "Name", "Date", "Description", "Reference", "Amount", "VAT", "Total", "Actions", "Done"
            ])
    # Ensure actions is always a list
    actions = tx.actions if tx.actions is not None else []
    if tx.type == 'received':
        sheet.append([
            tx.name, tx.date.isoformat(), tx.amount, tx.description, tx.method, ','.join(actions), tx.done
        ])
    else:
        sheet.append([
            tx.name, tx.date.isoformat(), tx.description, tx.reference, tx.amount, tx.vat, tx.total, ','.join(actions), tx.done
        ])
    wb.save(EXCEL_FILE)


def read_transactions_from_excel(tx_type: str):
    wb = get_or_create_workbook()
    sheet = wb[SHEET_NAMES[tx_type]]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = rows[0]
    transactions = []
    for row in rows[1:]:
        tx = dict(zip(headers, row))
        if tx_type == 'received':
            tx['Actions'] = [a for a in (tx.get('Actions') or '').split(',') if a]
        else:
            if tx.get('Actions'):
                tx['Actions'] = [a for a in tx['Actions'].split(',') if a]
            else:
                tx['Actions'] = []
        transactions.append(tx)
    return transactions


def update_transaction_in_excel(tx_type: str, idx: int, updated: dict):
    wb = get_or_create_workbook()
    sheet = wb[SHEET_NAMES[tx_type]]
    rows = list(sheet.iter_rows(values_only=False))
    if idx + 2 > len(rows):
        raise HTTPException(status_code=404, detail="Transaction not found.")
    headers = [cell.value for cell in rows[0]]
    # Map camelCase keys to Excel header names
    key_map = {
        'name': 'Name',
        'date': 'Date',
        'description': 'Description',
        'reference': 'Reference',
        'amount': 'Amount',
        'vat': 'VAT',
        'total': 'Total',
        'actions': 'Actions',
        'done': 'Done',
        'method': 'Method',
        'notes': 'Notes',
    }
    # Build a dict with Excel header keys
    excel_update = {}
    for k, v in updated.items():
        header = key_map.get(k, k)
        excel_update[header] = v
    for col, key in enumerate(headers):
        value = excel_update.get(key, rows[idx+1][col].value)
        if key == 'Actions' and isinstance(value, list):
            value = ','.join(value)
        rows[idx+1][col].value = value
    wb.save(EXCEL_FILE)


@app.post("/transaction")
def add_transaction(tx: Transaction):
    if tx.type not in SHEET_NAMES:
        raise HTTPException(status_code=400, detail="Invalid transaction type.")
    save_transaction_to_excel(tx)
    return {"message": "Transaction saved."}


@app.get("/transactions/{tx_type}")
def list_transactions(tx_type: str):
    if tx_type not in SHEET_NAMES:
        raise HTTPException(status_code=400, detail="Invalid transaction type.")
    txs = read_transactions_from_excel(tx_type)
    return JSONResponse(content=txs)


@app.put("/transactions/{tx_type}/{idx}")
def update_transaction(tx_type: str, idx: int, updated: dict = Body(...)):
    if tx_type not in SHEET_NAMES:
        raise HTTPException(status_code=400, detail="Invalid transaction type.")
    update_transaction_in_excel(tx_type, idx, updated)
    return {"message": "Transaction updated."}


@app.delete("/transactions/{tx_type}/{idx}")
def delete_transaction(tx_type: str, idx: int):
    if tx_type not in SHEET_NAMES:
        raise HTTPException(status_code=400, detail="Invalid transaction type.")
    wb = get_or_create_workbook()
    sheet = wb[SHEET_NAMES[tx_type]]
    rows = list(sheet.iter_rows(values_only=False))
    if idx + 2 > len(rows):
        raise HTTPException(status_code=404, detail="Transaction not found.")
    sheet.delete_rows(idx + 2)
    wb.save(EXCEL_FILE)
    return {"message": "Transaction deleted."}
